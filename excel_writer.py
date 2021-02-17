from constants import IMG_PROPERTY, EXCEL_FOLDER, EXCEL_WORKBOOK
import xlsxwriter
import openpyxl
import os


class ExcelWriter:
    def __init__(self, dictionaries):
        self.workbook = EXCEL_WORKBOOK
        self.dictionaries = dictionaries

    def write_excel_files(self):
        if not os.path.exists(self.workbook):
            self.create_workbook()

        workbook = openpyxl.load_workbook(filename=self.workbook)
        worksheet = workbook['Sheet1']
        if self.dictionaries:
            self.write_headers(worksheet)
            row = self.get_last_row() if os.path.exists(self.workbook) else 1
            for dictionary in self.dictionaries:
                self.write_to_worksheet(worksheet, row, dictionary)
                row += 1
            workbook.save(self.workbook)

    def create_workbook(self):
        workbook = xlsxwriter.Workbook(self.workbook)
        worksheet = workbook.add_worksheet()
        workbook.close()

    def write_headers(self, worksheet):
        col = 1
        headers = self.dictionaries[0].keys()
        for header in headers:
            worksheet.cell(row=1, column=col, value=header.upper())
            col += 1

    def get_last_row(self):
        workbook = openpyxl.load_workbook(filename=self.workbook)
        worksheet = workbook['Sheet1']
        return worksheet.max_row + 1

    def write_to_worksheet(self, worksheet, row, dictionary):
        col = 1
        for key, value in dictionary.items():
            if key == IMG_PROPERTY:
                worksheet.cell(row=row, column=col).hyperlink = value
            else:
                worksheet.cell(row=row, column=col, value=value)
            col+=1
            
